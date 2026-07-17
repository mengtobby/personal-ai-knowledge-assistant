"""Spreadsheet extraction: one segment per sheet, as CSV text.

.xlsx via openpyxl, legacy .xls via xlrd.
"""

import csv
from io import BytesIO, StringIO
from typing import Any, Iterable

from ..schemas import Segment
from .types import ParsedDocument, ParseError


def _rows_to_csv(rows: Iterable[Iterable[Any]]) -> str:
    out = StringIO()
    writer = csv.writer(out, lineterminator="\n")
    for row in rows:
        writer.writerow(["" if cell is None else cell for cell in row])
    return out.getvalue().strip()


def _xlsx_sheets(data: bytes) -> list[tuple[str, str]]:
    from openpyxl import load_workbook

    workbook = load_workbook(BytesIO(data), read_only=True, data_only=True)
    try:
        return [
            (sheet.title, _rows_to_csv(sheet.iter_rows(values_only=True)))
            for sheet in workbook.worksheets
        ]
    finally:
        workbook.close()


def _xls_sheets(data: bytes) -> list[tuple[str, str]]:
    import xlrd

    workbook = xlrd.open_workbook(file_contents=data)
    return [
        (sheet.name, _rows_to_csv(sheet.row_values(i) for i in range(sheet.nrows)))
        for sheet in workbook.sheets()
    ]


def parse_workbook(data: bytes, extension: str) -> ParsedDocument:
    try:
        sheets = _xls_sheets(data) if extension == ".xls" else _xlsx_sheets(data)
    except Exception as err:
        raise ParseError(f"Could not read this workbook: {err}") from err

    segments = [
        Segment(text=text, location=f'sheet "{name}"') for name, text in sheets if text
    ]
    if not segments:
        raise ParseError("Workbook contains no data.")
    return ParsedDocument(segments=segments)
